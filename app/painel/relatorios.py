"""Geração de relatórios PDF e Excel para o painel gerencial."""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
)

# ── Paleta de cores ───────────────────────────────────────────────────────────

_AZUL      = colors.HexColor('#0d6efd')
_AZUL_CLARO = colors.HexColor('#dbeafe')
_CINZA     = colors.HexColor('#6c757d')
_CINZA_CLARO = colors.HexColor('#f8f9fa')
_VERMELHO  = colors.HexColor('#dc3545')
_AMARELO   = colors.HexColor('#ffc107')
_VERDE     = colors.HexColor('#198754')
_BRANCO    = colors.white

_STATUS_PT = {
    'concluido':  'Concluído',
    'programado': 'Programado',
    'pendente':   'Pendente',
    'atencao':    'Atenção',
    'vencido':    'Vencido',
    'sem_periodo': '—',
}

_STATUS_COR = {
    'concluido':  _VERDE,
    'programado': _AZUL,
    'pendente':   _CINZA,
    'atencao':    _AMARELO,
    'vencido':    _VERMELHO,
    'sem_periodo': _CINZA,
}

# ── Helpers internos ──────────────────────────────────────────────────────────

def _fmt(d):
    """Formata date como dd/mm/AAAA, ou '—' se None."""
    return d.strftime('%d/%m/%Y') if d else '—'


def _header_footer(canvas, doc):
    """Cabeçalho e rodapé em todas as páginas."""
    canvas.saveState()
    w, h = doc.pagesize

    # Linha superior azul
    canvas.setFillColor(_AZUL)
    canvas.rect(0, h - 1.2 * cm, w, 1.2 * cm, fill=1, stroke=0)

    # Título no cabeçalho
    canvas.setFillColor(_BRANCO)
    canvas.setFont('Helvetica-Bold', 11)
    canvas.drawString(1 * cm, h - 0.85 * cm, 'Pellissari Fluidez Digital')

    canvas.setFont('Helvetica', 9)
    canvas.drawRightString(w - 1 * cm, h - 0.85 * cm,
                           f'Gerado em {_fmt(date.today())}')

    # Rodapé
    canvas.setFillColor(_CINZA)
    canvas.setFont('Helvetica', 8)
    canvas.drawString(1 * cm, 0.7 * cm,
                      'Pellissari Fluidez Digital — Relatório de Férias')
    canvas.drawRightString(w - 1 * cm, 0.7 * cm,
                           f'Página {doc.page}')

    # Linha inferior
    canvas.setStrokeColor(_CINZA)
    canvas.setLineWidth(0.3)
    canvas.line(1 * cm, 1.1 * cm, w - 1 * cm, 1.1 * cm)

    canvas.restoreState()


def _doc_base(buf, titulo, landscape_mode=True):
    """Cria SimpleDocTemplate com margens e cabeçalho padrão."""
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.6 * cm,
        title=titulo,
        author='Pellissari Fluidez Digital',
    )
    return doc


def _subtitulo_style():
    styles = getSampleStyleSheet()
    return ParagraphStyle(
        'Subtitulo',
        parent=styles['Normal'],
        fontSize=9,
        textColor=_CINZA,
        spaceAfter=6,
    )


# ── PDF: Status de Férias ─────────────────────────────────────────────────────

def gerar_pdf_ferias(colab_rows, filtro_equipe='todas'):
    """Retorna bytes de um PDF com tabela de status de férias por colaborador."""
    buf = io.BytesIO()
    doc = _doc_base(buf, 'Relatório de Férias')

    styles = getSampleStyleSheet()
    story = []

    # Subtítulo
    equipe_label = filtro_equipe if filtro_equipe != 'todas' else 'Todas as equipes'
    story.append(Paragraph(
        f'<b>Status de Férias por Colaborador</b> &nbsp;&nbsp;'
        f'<font color="#6c757d" size="9">Equipe: {equipe_label}</font>',
        ParagraphStyle('T', parent=styles['Normal'], fontSize=13, spaceAfter=8),
    ))
    story.append(Spacer(1, 0.3 * cm))

    # Cabeçalho da tabela
    cabecalho = ['Colaborador', 'Equipe', 'Função',
                 'Período aquisitivo', 'Dias\nrestantes',
                 'Limite de saída', 'Status']

    dados = [cabecalho]
    estilos_linha = []

    for i, row in enumerate(colab_rows, start=1):
        pp = row['periodo_princ']
        if pp:
            periodo_str = (f"{_fmt(pp['periodo'].data_inicio)}\n"
                           f"a {_fmt(pp['periodo'].data_fim)}")
            dias_str = str(pp['restantes']) if pp['restantes'] > 0 else '✓'
            limite_str = _fmt(pp['periodo'].data_limite_saida)
        else:
            periodo_str = '—'
            dias_str = '—'
            limite_str = '—'

        status = row['status']
        dados.append([
            row['colaborador'].nome,
            row['colaborador'].equipe.nome,
            row['colaborador'].funcao,
            periodo_str,
            dias_str,
            limite_str,
            _STATUS_PT.get(status, status),
        ])

        # Cor de fundo por status
        if status == 'vencido':
            bg = colors.HexColor('#fde8e8')
        elif status == 'atencao':
            bg = colors.HexColor('#fff9e6')
        else:
            bg = _CINZA_CLARO if i % 2 == 0 else _BRANCO
        estilos_linha.append(('BACKGROUND', (0, i), (-1, i), bg))

    col_widths = [5.5*cm, 3*cm, 3.5*cm, 4*cm, 1.8*cm, 3.2*cm, 2.5*cm]

    tabela = Table(dados, colWidths=col_widths, repeatRows=1)
    tabela.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND',   (0, 0), (-1, 0), _AZUL),
        ('TEXTCOLOR',    (0, 0), (-1, 0), _BRANCO),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0), 8),
        ('ALIGN',        (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',       (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING',(0, 0), (-1, 0), 7),
        ('TOPPADDING',   (0, 0), (-1, 0), 7),
        # Corpo
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
        ('VALIGN',       (0, 1), (-1, -1), 'MIDDLE'),
        ('ALIGN',        (4, 1), (4, -1), 'CENTER'),
        ('ALIGN',        (6, 1), (6, -1), 'CENTER'),
        ('TOPPADDING',   (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 1), (-1, -1), 4),
        # Grade
        ('GRID',         (0, 0), (-1, -1), 0.3, colors.HexColor('#dee2e6')),
        ('LINEBELOW',    (0, 0), (-1, 0), 1, _AZUL),
        *estilos_linha,
    ]))

    story.append(tabela)
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        f'Total: {len(colab_rows)} colaborador(es)',
        ParagraphStyle('Rod', parent=styles['Normal'], fontSize=8, textColor=_CINZA),
    ))

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buf.seek(0)
    return buf.read()


# ── PDF: Alertas ──────────────────────────────────────────────────────────────

def gerar_pdf_alertas(alertas, filtro_equipe='todas'):
    """Retorna bytes de um PDF somente com colaboradores em risco."""
    buf = io.BytesIO()
    doc = _doc_base(buf, 'Alertas de Férias', landscape_mode=False)

    styles = getSampleStyleSheet()
    sub = _subtitulo_style()
    story = []

    equipe_label = filtro_equipe if filtro_equipe != 'todas' else 'Todas as equipes'
    story.append(Paragraph(
        f'<b>Alertas de Férias</b> &nbsp;&nbsp;'
        f'<font color="#6c757d" size="9">Equipe: {equipe_label}</font>',
        ParagraphStyle('T', parent=styles['Normal'], fontSize=13, spaceAfter=8),
    ))

    secoes = [
        ('Vencidos',        alertas.get('vencidos', []),  _VERMELHO),
        ('Atenção ≤ 30 d',  alertas.get('dias_30', []),   _AMARELO),
        ('Atenção 31–60 d', alertas.get('dias_60', []),   colors.HexColor('#fd7e14')),
        ('No radar 61–90 d',alertas.get('dias_90', []),   _CINZA),
    ]

    cabecalho = ['Colaborador', 'Equipe', 'Função',
                 'Dias restantes', 'Limite de saída', 'Dias p/ vencer']

    for titulo, items, cor in secoes:
        if not items:
            continue

        story.append(Spacer(1, 0.4 * cm))
        story.append(Paragraph(f'<b>{titulo}</b>', sub))

        dados = [cabecalho]
        for a in items:
            dados.append([
                a['colaborador'].nome,
                a['colaborador'].equipe.nome,
                a['colaborador'].funcao,
                str(a['restantes']),
                _fmt(a['periodo'].data_limite_saida),
                str(a['dias_limite']),
            ])

        col_widths = [5*cm, 3*cm, 3.5*cm, 2.5*cm, 3*cm, 2.5*cm]
        tabela = Table(dados, colWidths=col_widths, repeatRows=1)
        tabela.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0), cor),
            ('TEXTCOLOR',    (0, 0), (-1, 0), _BRANCO),
            ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0), 8),
            ('ALIGN',        (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE',     (0, 1), (-1, -1), 8),
            ('ALIGN',        (3, 1), (5, -1), 'CENTER'),
            ('TOPPADDING',   (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
            ('GRID',         (0, 0), (-1, -1), 0.3, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [_BRANCO, _CINZA_CLARO]),
        ]))
        story.append(tabela)

    if not any(items for _, items, _ in secoes):
        story.append(Spacer(1, 1 * cm))
        story.append(Paragraph('Nenhum alerta no momento.', sub))

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buf.seek(0)
    return buf.read()


# ── Excel ─────────────────────────────────────────────────────────────────────

_XL_AZUL       = 'FF0D6EFD'
_XL_AZUL_CLARO = 'FFDBEAFE'
_XL_CINZA      = 'FFF8F9FA'
_XL_VERMELHO   = 'FFFDE8E8'
_XL_AMARELO    = 'FFFFF9E6'


def _xl_header_style():
    return Font(bold=True, color='FFFFFFFF', size=10)


def _xl_header_fill():
    return PatternFill('solid', fgColor=_XL_AZUL)


def _xl_center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _xl_wrap():
    return Alignment(wrap_text=True, vertical='top')


def gerar_excel_ferias(colab_rows, alertas, filtro_equipe='todas'):
    """Retorna bytes de um Excel com aba de status e aba de alertas."""
    wb = Workbook()

    # ── Aba 1: Status Férias ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Status Férias'

    equipe_label = filtro_equipe if filtro_equipe != 'todas' else 'Todas as equipes'
    ws['A1'] = 'Pellissari Fluidez Digital — Relatório de Férias'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A2'] = f'Equipe: {equipe_label}   |   Gerado em: {_fmt(date.today())}'
    ws['A2'].font = Font(color='FF6C757D', size=9)
    ws.append([])

    cabecalho = ['Colaborador', 'Equipe', 'Função',
                 'Início período', 'Fim período',
                 'Dias restantes', 'Limite de saída', 'Status']
    ws.append(cabecalho)

    hdr_row = ws.max_row
    for col in range(1, len(cabecalho) + 1):
        cell = ws.cell(row=hdr_row, column=col)
        cell.font = _xl_header_style()
        cell.fill = _xl_header_fill()
        cell.alignment = _xl_center()

    for i, row in enumerate(colab_rows):
        pp = row['periodo_princ']
        if pp:
            inicio = pp['periodo'].data_inicio
            fim = pp['periodo'].data_fim
            dias = pp['restantes'] if pp['restantes'] > 0 else 0
            limite = pp['periodo'].data_limite_saida
        else:
            inicio = fim = limite = None
            dias = None

        status = row['status']
        ws.append([
            row['colaborador'].nome,
            row['colaborador'].equipe.nome,
            row['colaborador'].funcao,
            inicio,
            fim,
            dias,
            limite,
            _STATUS_PT.get(status, status),
        ])

        data_row = ws.max_row
        # Formatar colunas de data
        for col in (4, 5, 7):
            ws.cell(row=data_row, column=col).number_format = 'DD/MM/YYYY'
        ws.cell(row=data_row, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=data_row, column=8).alignment = Alignment(horizontal='center')

        # Cor de fundo por status
        if status == 'vencido':
            fill = PatternFill('solid', fgColor='FFFDE8E8')
        elif status == 'atencao':
            fill = PatternFill('solid', fgColor='FFFFF9E6')
        elif i % 2 == 0:
            fill = PatternFill('solid', fgColor='FFF8F9FA')
        else:
            fill = None

        if fill:
            for col in range(1, len(cabecalho) + 1):
                ws.cell(row=data_row, column=col).fill = fill

    # Larguras de coluna
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14

    ws.freeze_panes = 'A5'

    # ── Aba 2: Alertas ────────────────────────────────────────────────────────
    wa = wb.create_sheet('Alertas')
    wa['A1'] = 'Pellissari Fluidez Digital — Alertas de Férias'
    wa['A1'].font = Font(bold=True, size=12)
    wa['A2'] = f'Equipe: {equipe_label}   |   Gerado em: {_fmt(date.today())}'
    wa['A2'].font = Font(color='FF6C757D', size=9)
    wa.append([])

    secoes = [
        ('Vencidos',         alertas.get('vencidos', []),  'FFFDE8E8', 'FFDC3545'),
        ('Atenção ≤ 30 dias',alertas.get('dias_30', []),   'FFFFF9E6', 'FFFFC107'),
        ('Atenção 31–60 d',  alertas.get('dias_60', []),   'FFFFF3E0', 'FFFD7E14'),
        ('No radar 61–90 d', alertas.get('dias_90', []),   'FFF8F9FA', 'FF6C757D'),
    ]

    cab_alertas = ['Colaborador', 'Equipe', 'Função',
                   'Dias restantes', 'Limite de saída', 'Dias p/ vencer']

    tem_dados = False
    for titulo, items, bg_fill, hdr_color in secoes:
        if not items:
            continue
        tem_dados = True

        wa.append([titulo])
        sec_row = wa.max_row
        wa.cell(row=sec_row, column=1).font = Font(bold=True, size=10,
                                                    color=hdr_color)
        wa.append(cab_alertas)
        hdr_row = wa.max_row
        for col in range(1, len(cab_alertas) + 1):
            cell = wa.cell(row=hdr_row, column=col)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill('solid', fgColor=hdr_color)
            cell.alignment = _xl_center()

        for j, a in enumerate(items):
            wa.append([
                a['colaborador'].nome,
                a['colaborador'].equipe.nome,
                a['colaborador'].funcao,
                a['restantes'],
                a['periodo'].data_limite_saida,
                a['dias_limite'],
            ])
            data_row = wa.max_row
            wa.cell(row=data_row, column=5).number_format = 'DD/MM/YYYY'
            wa.cell(row=data_row, column=4).alignment = Alignment(horizontal='center')
            wa.cell(row=data_row, column=6).alignment = Alignment(horizontal='center')
            if j % 2 == 0:
                for col in range(1, len(cab_alertas) + 1):
                    wa.cell(row=data_row, column=col).fill = PatternFill(
                        'solid', fgColor=bg_fill)

        wa.append([])

    if not tem_dados:
        wa.append(['Nenhum alerta no momento.'])

    for col_letter, width in zip('ABCDEF', [28, 16, 20, 14, 16, 14]):
        wa.column_dimensions[col_letter].width = width

    wa.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
